from datetime import datetime
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.reporting.service import ReportDataset


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_BORDER = Border(bottom=Side(style="thin", color="9EADBA"))


def style_table(sheet, widths: list[float], freeze: str = "A2") -> None:
    sheet.freeze_panes = freeze
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[chr(64 + index)].width = width
    sheet.auto_filter.ref = sheet.dimensions


def build_workbook(dataset: ReportDataset, exported_at: datetime) -> Workbook:
    workbook = Workbook()
    overview = workbook.active
    overview.title = "Tổng quan"
    overview.append(["Facebook Page Report", None])
    overview.append(["Xuất lúc", exported_at])
    overview.append(["Khoảng báo cáo", f"{dataset.start_date.isoformat()} — {dataset.end_date.isoformat()}"])
    overview.append(["Page", dataset.page_name])
    overview.append(["Page ID", dataset.external_page_id])
    overview.append(["Followers", dataset.summary.current_followers])
    overview.append(["Category", dataset.category])
    overview.append(["Link", dataset.link])
    overview["A1"].font = Font(size=15, bold=True, color="1F4E78")
    overview.column_dimensions["A"].width = 26
    overview.column_dimensions["B"].width = 60

    posts = workbook.create_sheet("Posts")
    posts.append(["Created", "Message", "Media", "Reactions", "Like", "Love", "Haha", "Wow", "Sad", "Angry", "Care", "Comments", "Shares", "Engagement", "Permalink"])
    for row in dataset.posts:
        posts.append([row.created, row.message, row.media, row.reactions, row.like, row.love, row.haha, row.wow, row.sad, row.angry, row.care, row.comments, row.shares, row.engagement, row.permalink])
    style_table(posts, [20, 70, 14, 12, 9, 9, 9, 9, 9, 9, 9, 12, 10, 13, 55])

    insights = workbook.create_sheet("Insights")
    insights.append(["Date", "page_post_engagements", "page_views_total", "page_video_views", "page_follows", "page_daily_follows", "page_daily_follows_unique", "page_daily_unfollows_unique"])
    for row in dataset.daily_insights:
        insights.append([row.metric_date, row.post_engagements, row.page_views, row.video_views, row.followers_total, row.daily_new_followers, row.daily_new_followers, None])
    style_table(insights, [15, 24, 20, 20, 18, 22, 28, 30], "B2")

    videos = workbook.create_sheet("Videos")
    videos.append(["Title", "Created", "Length (s)", "Views", "Avg Watch Time", "Complete Views", "Permalink"])
    for row in dataset.videos:
        videos.append([row["title"], row["created"], row["length_seconds"], row["views"], row["average_watch_time"], row["complete_views"], row["permalink"]])
    style_table(videos, [60, 20, 14, 12, 20, 18, 55])

    formats = workbook.create_sheet("Định dạng")
    formats.append(["Định dạng", "Số bài", "Engagement TB", "Tổng Engagement", "Reactions", "Comments", "Shares"])
    for row in dataset.formats:
        formats.append([row.media, row.posts, row.average_engagement, row.total_engagement, row.reactions, row.comments, row.shares])
    style_table(formats, [18, 12, 18, 20, 14, 14, 12])

    posting = workbook.create_sheet("Giờ đăng")
    posting.append(["Giờ VN \\ Thứ", "T2", "T3", "T4", "T5", "T6", "T7", "CN"])
    for hour, values in enumerate(dataset.posting_hours):
        posting.append([f"{hour:02d}h", *values])
    style_table(posting, [18, 10, 10, 10, 10, 10, 10, 10])

    comments = workbook.create_sheet("Comments")
    comments.append(["Thời gian", "Bài viết", "Nội dung comment", "Người bình luận", "Likes", "Replies"])
    for row in dataset.comments_detail:
        comments.append([row["created"], row["post_excerpt"], row["message"], row["author"], row["likes"], row["replies"]])
    style_table(comments, [20, 40, 70, 24, 10, 10])

    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, datetime):
                    if cell.value.tzinfo is not None:
                        cell.value = cell.value.astimezone(ZoneInfo("Asia/Bangkok")).replace(tzinfo=None)
                    cell.number_format = "yyyy-mm-dd hh:mm:ss"
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    return workbook
