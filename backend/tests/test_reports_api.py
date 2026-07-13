from datetime import UTC, datetime
from contextlib import contextmanager
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session, sessionmaker
from sqlalchemy.pool import StaticPool

from app.dependencies import get_current_admin, get_session
from app.main import create_app
from app.models import AdminUser, Base, Page, Post


@contextmanager
def build_report_client():
    engine = create_engine("sqlite+pysqlite:///:memory:", connect_args={"check_same_thread": False}, poolclass=StaticPool)
    Base.metadata.create_all(engine)
    factory = sessionmaker(bind=engine, expire_on_commit=False)
    with factory() as db:
        page = Page(external_page_id="1125132200689307", display_name="HTV")
        db.add(page); db.flush()
        db.add(Post(page_id=page.id, external_post_id="p1", published_at=datetime(2026, 7, 1, 1, tzinfo=UTC), media_type="photo", reactions=3, comment_count=2, share_count=1))
        db.commit(); page_id = page.id
    app = create_app()

    def override_session():
        with factory() as db:
            yield db

    app.dependency_overrides[get_session] = override_session
    app.dependency_overrides[get_current_admin] = lambda: AdminUser(id=1, username="admin", password_hash="x")
    try:
        with TestClient(app) as client:
            yield client, page_id
    finally:
        engine.dispose()


def test_report_api_returns_shared_kpis():
    with build_report_client() as (client, page_id):
        response = client.get(f"/api/reports/{page_id}?from=2026-07-01&to=2026-07-02")

    assert response.status_code == 200
    assert response.json()["summary"]["total_engagement"] == 6


def test_excel_endpoint_streams_filtered_workbook():
    with build_report_client() as (client, page_id):
        response = client.get(f"/api/reports/{page_id}/excel?from=2026-07-01&to=2026-07-02")

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "facebook-report_1125132200689307_2026-07-01_2026-07-02.xlsx" in response.headers["content-disposition"]
    assert load_workbook(BytesIO(response.content)).sheetnames[0] == "Tổng quan"
